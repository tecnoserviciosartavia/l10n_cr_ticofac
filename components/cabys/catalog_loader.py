import hashlib
import logging
from datetime import datetime, timezone
from pathlib import Path

from odoo.tools import file_path
from openpyxl import load_workbook

_logger = logging.getLogger(__name__)

CATALOG_RESOURCE = "l10n_cr_ticofac/components/cabys/data/Catalogo-CABYS-vigente.xlsx"
CATEGORY_WIDTHS = (1, 2, 3, 4, 5, 7, 9, 11)
PRODUCT_WIDTH = 13
BATCH_SIZE = 1000


def _text(value):
    if value is None:
        return ""
    return str(value).strip()


def _code(value, width):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip().replace(".0", "").zfill(width)


def _tax_percentage(value):
    if value in (None, ""):
        return 0.0
    if isinstance(value, str):
        normalized = value.strip().replace("%", "").replace(",", ".")
        if normalized.lower() in {"exento", "na", "n-a", ""}:
            return 0.0
        number = float(normalized)
        return number if "%" in value else (number * 100 if number <= 1 else number)
    number = float(value)
    return number * 100 if number <= 1 else number


def _read_catalog(catalog_path):
    workbook = load_workbook(catalog_path, read_only=True, data_only=True)
    worksheet = workbook["Catálogo"]
    categories = {level: {} for level in range(1, 9)}
    products = {}

    for row in worksheet.iter_rows(min_row=3, values_only=True):
        category_codes = []
        for level, width in enumerate(CATEGORY_WIDTHS, start=1):
            code = _code(row[(level - 1) * 2], width)
            name = _text(row[(level - 1) * 2 + 1])
            if not code:
                break
            category_codes.append(code)
            categories[level][code] = {
                "name": name or code,
                "parent_code": category_codes[-2] if level > 1 else False,
            }
        if len(category_codes) != 8:
            continue

        product_code = _code(row[16], PRODUCT_WIDTH)
        if not product_code:
            continue
        products[product_code] = {
            "name": _text(row[17]) or product_code,
            "impuesto": _tax_percentage(row[18]),
            "category_code": category_codes[-1],
            "first_description": _text(row[19]),
            "second_description": _text(row[20]),
        }

    modified = workbook.properties.modified
    workbook.close()
    return categories, products, modified


def _chunks(values, size=BATCH_SIZE):
    for index in range(0, len(values), size):
        yield values[index:index + size]


def load_bundled_catalog(env, force=False):
    """Load or refresh the bundled official CABYS snapshot, keyed by CABYS code."""
    Product = env["cabys.producto"].sudo().with_context(tracking_disable=True)
    if not force and Product.search_count([]):
        _logger.info("CABYS initial load skipped because the catalog already contains records")
        return {"created": 0, "updated": 0, "total": Product.search_count([]), "skipped": True}

    catalog_path = Path(file_path(CATALOG_RESOURCE))
    digest = hashlib.sha256(catalog_path.read_bytes()).hexdigest()
    categories, products, modified = _read_catalog(catalog_path)
    category_ids = {}

    for level in range(1, 9):
        Model = env[f"cabys.categoria{level}"].sudo()
        existing = {record.codigo: record for record in Model.search([])}
        create_values = []
        for code, data in categories[level].items():
            values = {"codigo": code, "name": data["name"]}
            if level > 1:
                values[f"cabys_categoria{level - 1}_id"] = category_ids[data["parent_code"]]
            record = existing.get(code)
            if record:
                changed = {}
                if record.name != values["name"]:
                    changed["name"] = values["name"]
                if level > 1:
                    parent_field = f"cabys_categoria{level - 1}_id"
                    if record[parent_field].id != values[parent_field]:
                        changed[parent_field] = values[parent_field]
                if changed:
                    record.write(changed)
                category_ids[code] = record.id
            else:
                create_values.append(values)
        for batch in _chunks(create_values):
            for record in Model.create(batch):
                category_ids[record.codigo] = record.id

    existing_products = {record.codigo: record for record in Product.search([])}
    create_values = []
    updated = 0
    for code, data in products.items():
        values = {
            "codigo": code,
            "name": data["name"],
            "impuesto": data["impuesto"],
            "cabys_categoria8_id": category_ids[data["category_code"]],
            "first_description": data["first_description"],
            "second_description": data["second_description"],
        }
        record = existing_products.get(code)
        if not record:
            create_values.append(values)
            continue
        changed = {}
        for field_name in ("name", "impuesto", "first_description", "second_description"):
            if (record[field_name] or "") != (values[field_name] or ""):
                changed[field_name] = values[field_name]
        if record.cabys_categoria8_id.id != values["cabys_categoria8_id"]:
            changed["cabys_categoria8_id"] = values["cabys_categoria8_id"]
        if changed:
            record.write(changed)
            updated += 1

    created = 0
    for batch in _chunks(create_values):
        Product.create(batch)
        created += len(batch)
        _logger.info("CABYS initial load progress: %s/%s products", created, len(create_values))

    params = env["ir.config_parameter"].sudo()
    params.set_param("l10n_cr_ticofac.cabys.sha256", digest)
    params.set_param("l10n_cr_ticofac.cabys.products", len(products))
    params.set_param("l10n_cr_ticofac.cabys.source_modified", modified.isoformat() if modified else "")
    params.set_param("l10n_cr_ticofac.cabys.loaded_at", datetime.now(timezone.utc).isoformat())
    _logger.info("CABYS catalog loaded: %s created, %s updated, %s total", created, updated, len(products))
    return {"created": created, "updated": updated, "total": len(products), "skipped": False}
